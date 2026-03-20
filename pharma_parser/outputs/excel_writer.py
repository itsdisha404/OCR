"""
excel_writer.py
───────────────
Write invoice data to a master Excel workbook with daily sheets.
Each row = one line item with invoice header fields repeated.
"""

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# Column order in Excel
EXCEL_COLUMNS = [
    "invoice_no", "invoice_date", "vendor_name", "buyer_name", "vendor_id",
    "product_name", "hsn_code", "batch_no", "expiry_date", "pack_size",
    "qty_billed", "qty_free", "mrp", "ptr", "pts", "rate",
    "discount_pct", "discount_value", "taxable_value",
    "cgst_rate", "cgst_amount", "sgst_rate", "sgst_amount", "igst_amount",
    "total_amount", "grand_total", "confidence", "flags", "source_file",
]

# Styling
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
LOW_CONF_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def append_to_excel(invoice: dict, excel_path) -> None:
    """
    Append a single invoice's line items to the master Excel workbook.
    Creates a new sheet per day (named YYYY-MM-DD).
    Rows with LOW confidence get red background for easy review.

    Args:
        invoice: Full invoice dict with header, line_items, summary.
        excel_path: Path to the Excel file.
    """
    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    # Load or create workbook
    if excel_path.exists():
        wb = load_workbook(str(excel_path))
    else:
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Get or create today's sheet
    sheet_name = str(date.today())
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        # Write header row with styling
        for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        # Set column widths
        for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = 15

    # Extract header values
    h = invoice.get("header", {})
    s = invoice.get("summary", {})

    # Write one row per line item
    for item in invoice.get("line_items", []):
        row_data = [
            h.get("invoice_no", ""),
            h.get("invoice_date", ""),
            h.get("vendor_name", ""),
            h.get("buyer_name", ""),
            h.get("vendor_id", ""),
            item.get("product_name", ""),
            item.get("hsn_code", ""),
            item.get("batch_no", ""),
            item.get("expiry_date", ""),
            item.get("pack_size", ""),
            item.get("qty_billed", 0),
            item.get("qty_free", 0),
            item.get("mrp", 0),
            item.get("ptr", 0),
            item.get("pts", 0),
            item.get("rate", 0),
            item.get("discount_pct", 0),
            item.get("discount_value", 0),
            item.get("taxable_value", 0),
            item.get("cgst_rate", 0),
            item.get("cgst_amount", 0),
            item.get("sgst_rate", 0),
            item.get("sgst_amount", 0),
            item.get("igst_amount", 0),
            item.get("total_amount", 0),
            s.get("grand_total", 0),
            item.get("confidence", ""),
            "|".join(item.get("flags", [])),
            h.get("source_file", ""),
        ]

        row_num = ws.max_row + 1
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            # Highlight LOW confidence rows
            if item.get("confidence") == "LOW":
                cell.fill = LOW_CONF_FILL

    wb.save(str(excel_path))


def append_multiple_to_excel(invoices: list, excel_path) -> None:
    """Append multiple invoices to the same Excel file."""
    for inv in invoices:
        append_to_excel(inv, excel_path)
